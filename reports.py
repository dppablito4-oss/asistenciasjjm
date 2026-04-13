from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional, Tuple

import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

from database import DB_PATH, get_connection


def _to_date(value: Optional[str]) -> date:
    if isinstance(value, date):
        return value
    if not value:
        return date.today()
    return datetime.strptime(value, "%Y-%m-%d").date()


def _period_bounds(
    period: str,
    ref_date: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> Tuple[date, date]:
    period = (period or "day").lower()
    ref = _to_date(ref_date)

    if period == "day":
        return ref, ref
    if period == "week":
        monday = ref - timedelta(days=ref.weekday())
        sunday = monday + timedelta(days=6)
        return monday, sunday
    if period == "month":
        first_day = ref.replace(day=1)
        if first_day.month == 12:
            next_month = first_day.replace(year=first_day.year + 1, month=1)
        else:
            next_month = first_day.replace(month=first_day.month + 1)
        last_day = next_month - timedelta(days=1)
        return first_day, last_day
    if period == "range":
        if not start_date or not end_date:
            raise ValueError("Range period requires start_date and end_date.")
        start = _to_date(start_date)
        end = _to_date(end_date)
        if start > end:
            raise ValueError("start_date must be <= end_date")
        return start, end

    raise ValueError("Invalid period. Use day, week, month, or range.")


def _build_students_query(grado, seccion, genero, cargo):
    conditions = ["activo = 1"]
    params = []

    if grado and str(grado).lower() != "todos":
        conditions.append("grado = ?")
        params.append(int(grado))
    if seccion and str(seccion).lower() != "todos":
        conditions.append("seccion = ?")
        params.append(str(seccion).upper())
    if genero and str(genero).lower() != "todos":
        conditions.append("genero = ?")
        params.append(str(genero).upper())
    if cargo and str(cargo).lower() != "todos":
        conditions.append("cargo = ?")
        params.append(str(cargo))

    where_sql = ""
    if conditions:
        where_sql = " WHERE " + " AND ".join(conditions)

    sql = (
        "SELECT dni, nombres, apellidos, grado, seccion, genero, cargo "
        "FROM estudiantes"
        f"{where_sql} "
        "ORDER BY apellidos, nombres"
    )
    return sql, params


def generate_report(
    db_path: Path = DB_PATH,
    period: str = "day",
    ref_date: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    grado: Optional[str] = None,
    seccion: Optional[str] = None,
    genero: Optional[str] = None,
    cargo: Optional[str] = None,
    condition: str = "all",
) -> pd.DataFrame:
    start, end = _period_bounds(period, ref_date, start_date, end_date)

    with get_connection(db_path) as conn:
        students_sql, students_params = _build_students_query(grado, seccion, genero, cargo)
        df_students = pd.read_sql_query(students_sql, conn, params=students_params)

        if df_students.empty:
            return pd.DataFrame(
                columns=[
                    "fecha",
                    "hora",
                    "dni",
                    "nombres",
                    "apellidos",
                    "grado",
                    "seccion",
                    "genero",
                    "cargo",
                    "profesor_encargado",
                    "estado",
                ]
            )

        placeholders = ",".join("?" for _ in df_students["dni"].tolist())
        params = [start.isoformat(), end.isoformat(), *df_students["dni"].tolist()]
        attendance_sql = (
            "SELECT fecha, hora, estudiante_dni AS dni, profesor_encargado, estado "
            "FROM asistencia "
            "WHERE fecha BETWEEN ? AND ? "
            f"AND estudiante_dni IN ({placeholders})"
        )
        df_attendance = pd.read_sql_query(attendance_sql, conn, params=params)

    df_present = df_attendance.merge(df_students, how="left", on="dni")
    df_present = df_present[
        [
            "fecha",
            "hora",
            "dni",
            "nombres",
            "apellidos",
            "grado",
            "seccion",
            "genero",
            "cargo",
            "profesor_encargado",
            "estado",
        ]
    ]

    all_days = pd.date_range(start=start, end=end, freq="D").strftime("%Y-%m-%d")
    df_dates = pd.DataFrame({"fecha": list(all_days)})

    df_expected = df_students.assign(_k=1).merge(df_dates.assign(_k=1), on="_k").drop(columns=["_k"])
    present_keys = (
        df_present[["dni", "fecha"]]
        .drop_duplicates()
        .assign(_present=1)
    )
    df_absent = df_expected.merge(present_keys, how="left", on=["dni", "fecha"])
    df_absent = df_absent[df_absent["_present"].isna()].drop(columns=["_present"])
    df_absent["hora"] = "--:--:--"
    df_absent["profesor_encargado"] = ""
    df_absent["estado"] = "Falto"
    df_absent = df_absent[
        [
            "fecha",
            "hora",
            "dni",
            "nombres",
            "apellidos",
            "grado",
            "seccion",
            "genero",
            "cargo",
            "profesor_encargado",
            "estado",
        ]
    ]

    condition = (condition or "all").lower()
    if condition in {"present", "asistieron"}:
        final_df = df_present.copy()
    elif condition in {"absent", "faltaron"}:
        final_df = df_absent.copy()
    else:
        final_df = pd.concat([df_present, df_absent], ignore_index=True)

    final_df = final_df.sort_values(by=["fecha", "grado", "seccion", "apellidos", "nombres", "hora"])
    final_df.reset_index(drop=True, inplace=True)
    return final_df


def export_report_to_excel(
    df: pd.DataFrame,
    file_path: str,
    school_name: str = "Asistencia Escolar",
    logo_path: Optional[str] = None,
) -> None:
    if df.empty:
        raise ValueError("No data to export")

    df_to_export = df.copy()
    start_row = 4
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df_to_export.to_excel(writer, sheet_name="Reporte", index=False, startrow=start_row - 1)
        ws = writer.sheets["Reporte"]

        ws["A1"] = school_name or "Asistencia Escolar"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(vertical="center")
        ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(size=10)
        ws["A2"].alignment = Alignment(vertical="center")

        if logo_path:
            p = Path(logo_path)
            if p.exists():
                img = XLImage(str(p))
                img.width = 60
                img.height = 60
                ws.add_image(img, "J1")

        # Basic clean table formatting: header bold and auto width.
        for cell in ws[start_row]:
            cell.font = cell.font.copy(bold=True)

        for col in ws.iter_cols(min_row=start_row, max_row=ws.max_row):
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

        ws.row_dimensions[1].height = 34
        ws.row_dimensions[2].height = 20


def export_report_to_pdf(
    df: pd.DataFrame,
    file_path: str,
    school_name: str = "Asistencia Escolar",
    logo_path: Optional[str] = None,
    report_type: str = "General",
    generated_by: str = "",
) -> None:
    if df.empty:
        raise ValueError("No data to export")

    out = Path(file_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    c = canvas.Canvas(str(out), pagesize=A4)
    page_w, page_h = A4
    margin_x = 14 * mm
    margin_top = 14 * mm
    margin_bottom = 14 * mm

    cols = [
        ("fecha", 22 * mm),
        ("hora", 16 * mm),
        ("dni", 20 * mm),
        ("nombres", 33 * mm),
        ("apellidos", 33 * mm),
        ("grado", 12 * mm),
        ("seccion", 14 * mm),
        ("estado", 18 * mm),
    ]

    def _draw_header(page_no: int):
        y = page_h - margin_top
        c.setFillColor(colors.HexColor("#0f172a"))
        c.roundRect(margin_x, y - 16 * mm, page_w - 2 * margin_x, 16 * mm, 4, stroke=0, fill=1)

        if logo_path:
            p = Path(logo_path)
            if p.exists() and p.is_file():
                try:
                    c.drawImage(ImageReader(str(p)), margin_x + 3 * mm, y - 14.5 * mm, width=11 * mm, height=11 * mm, preserveAspectRatio=True, mask="auto")
                except Exception:
                    pass

        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margin_x + 17 * mm, y - 7 * mm, (school_name or "ASISTENCIA ESCOLAR")[:52])

        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.setFont("Helvetica", 8)
        c.drawRightString(page_w - margin_x - 3 * mm, y - 6 * mm, f"Reporte: {report_type}")
        c.drawRightString(page_w - margin_x - 3 * mm, y - 10 * mm, f"Generado: {stamp}")
        if generated_by:
            c.drawRightString(page_w - margin_x - 3 * mm, y - 14 * mm, f"Operador: {generated_by}")

        table_top = y - 20 * mm
        c.setFillColor(colors.HexColor("#9ca3af"))
        c.rect(margin_x, table_top - 6 * mm, page_w - 2 * margin_x, 6 * mm, stroke=0, fill=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 7)
        x = margin_x + 1.5 * mm
        for name, width in cols:
            c.drawString(x, table_top - 4.2 * mm, name.upper())
            x += width

        c.setFont("Helvetica", 7)
        return table_top - 8 * mm

    def _draw_page_number(page_no: int):
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 8)
        c.drawRightString(page_w - margin_x, margin_bottom - 4 * mm, f"Pagina {page_no}")

    page_no = 1
    y = _draw_header(page_no)
    row_h = 5.2 * mm
    max_y = margin_bottom + 4 * mm
    row_index = 0

    for _, row in df.iterrows():
        if y - row_h < max_y:
            _draw_page_number(page_no)
            c.showPage()
            page_no += 1
            y = _draw_header(page_no)

        x = margin_x + 1.5 * mm
        values = [
            str(row.get("fecha", "")),
            str(row.get("hora", "")),
            str(row.get("dni", "")),
            str(row.get("nombres", ""))[:24],
            str(row.get("apellidos", ""))[:24],
            str(row.get("grado", "")),
            str(row.get("seccion", "")),
            str(row.get("estado", "")),
        ]

        if row_index % 2 == 0:
            c.setFillColor(colors.HexColor("#f3f4f6"))
            c.rect(margin_x, y - row_h, page_w - 2 * margin_x, row_h, stroke=0, fill=1)

        for (col_name, width), value in zip(cols, values):
            c.setFillColor(colors.black)
            c.drawString(x, y - 3.8 * mm, value)
            x += width

        c.setStrokeColor(colors.HexColor("#4b5563"))
        c.setLineWidth(0.25)
        c.line(margin_x, y - row_h, page_w - margin_x, y - row_h)
        y -= row_h
        row_index += 1

    _draw_page_number(page_no)
    c.save()
